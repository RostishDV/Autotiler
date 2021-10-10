from openpyxl import load_workbook


class ExcelTileParser:
    def __init__(self, filename):
        self.wb = load_workbook(filename)
        self.rules_action = {}
        self.scripts_map = {}
        self.io_id_to_name = {}
        # print(self.wb.sheetnames)

    #rules_action = {
    #   field_name: {
    #       'rules': {
    #           rule_name_1: {
    #               'rule': rule,
    #               'visible': visible,
    #               'required': required,
    #               'read_only': read_only
    #           },
    #           rule_name_2: {
    #               'rule': rule,
    #               'visible': visible,
    #               'required': required,
    #               'read_only': read_only
    #           },
    #       },
    #       'full_name': full_field_name
    #   }
    # }
    def fill_id_to_names_list(self, sheet_name):
        ws = self.wb[sheet_name]
        id_col = 5
        field_name_col = 6
        for i in range(4, ws.max_row + 1):
            row = ws[i]
            field_id = row[id_col].value
            field_name = row[field_name_col].value
            self.io_id_to_name[field_id] = field_name

    def read_rules_actions(self, sheet_name):
        ws = self.wb[sheet_name]
        need_not_to_do = ['#Н/Д', '#N/A']
        header = ws[1]
        tile_field_col = 0
        politic_name_col = 0
        is_required_col = 0
        is_read_only_col = 0
        is_visible_col = 0
        i = 0
        for cell in header:
            if cell.value == 'ВПР живые поля с плитки/заголовка':
                tile_field_col = i
            if cell.value == 'Политика интерфейса пользователя':
                politic_name_col = i
            if cell.value == 'Видимое':
                is_visible_col = i
            if cell.value == 'Только для чтения':
                is_read_only_col = i
            if cell.value == 'Обязательный':
                is_required_col = i
            i += 1
        for i in range(3, ws.max_row + 1):
            row = ws[i]
            if row[tile_field_col].value in need_not_to_do:
                continue
            tile_field = row[tile_field_col].value
            visible = self.get_column_right_version(row, is_visible_col, 'Видимость')
            required = self.get_column_right_version(row, is_required_col, 'Обязательнотсь')
            enabled = self.get_column_right_version(row, is_read_only_col, 'редактируемость')
            info = {
                    'visible': visible, 
                    'required': required, 
                    'enabled': enabled
                }
            if tile_field in self.rules_action:
                self.rules_action[tile_field]['rules'][row[politic_name_col].value] = info
            else:
                self.rules_action[tile_field] = {}
                self.rules_action[tile_field]['rules'] = {}
                self.rules_action[tile_field]['rules'][row[politic_name_col].value] = info
    

    def get_column_right_version(self, row, column, property_name):
        not_touch = 'Не трогать'
        yes = 'Верно'
        no = 'Неверно'
        if row[column].value == not_touch:
            return ''
        if row[column].value == yes:
            if property_name == 'редактируемость':
                return f'!{property_name}'
            else:
                return property_name
        if row[column].value == no:
            if property_name == 'редактируемость':
                return property_name
            else:
                return f'!{property_name}'


    def read_rules(self, sheet_name):
        ws = self.wb[sheet_name]
        rule_col = 0
        rule_name_col = 0
        header = ws[1]
        i = 0
        for cell in header:
            if cell.value == 'Условия каталога':
                rule_col = i
            if cell.value == 'Краткое описание':
                rule_name_col = i
            i += 1

        for i in range(3, ws.max_row + 1):
            row = ws[i]
            rule = row[rule_col].value
            rule_name = row[rule_name_col].value
            self.add_in_rules_actions(rule_name, rule)


    def add_in_rules_actions(self, rule_name, rule):
        for field_name in self.rules_action.keys():
            if rule_name in self.rules_action[field_name]['rules']:
                if not rule == '':
                    rule = self.get_rule_right_version(rule)
                self.rules_action[field_name]['rules'][rule_name]['rule'] = rule
    
    
    def get_rule_right_version(self, rule):
        full_rule = rule
        full_rule = full_rule.replace('ORIO', 'OR\n\tIO')
        full_rule = full_rule.replace('NQIO', 'NQ\n\tIO')
        full_rule = full_rule.replace('^IO', '^\n\tIO')
        for field_id in self.io_id_to_name:
            full_rule = full_rule.replace(field_id, f'{self.io_id_to_name[field_id]} ')
        return full_rule

    def try_fill_full_field_name(self, sheet_name):
        ws = self.wb[sheet_name]
        field_name_col = 6
        full_field_name_col = 7
        header = ws[1]
        i = 0
        for i in range(3, ws.max_row + 1):
            row = ws[i]
            field = row[field_name_col].value
            if field in self.rules_action:
                full_name = row[full_field_name_col].value
                self.rules_action[field]['full_name'] = full_name


    def write_in_file(self):
        # self.print_rules_actions()
        with open('out/БП.txt', 'w', encoding='utf-8') as f:
            for field_name in self.rules_action.keys():
                for rule_name in self.rules_action[field_name]['rules']:
                    info = self.rules_action[field_name]['rules'][rule_name]
                    visible = info['visible']
                    required = info['required']
                    enabled = info['enabled']
                    rule = info['rule']
                    full_name = self.rules_action[field_name]['full_name'] if self.rules_action[field_name]['full_name'] else 'NOT_FOUND_IN_IO'
                    f.write(f'{rule_name}\n\n')
                    if rule != '':
                        f.write(f'\t{rule}\n\n')
                    f.write(f'\t{field_name} {full_name} \t\t{visible} {required} {enabled}\n\n')
                f.write('================================================\n\n')


    def fill_source_file (self, sc_cat_item_sheet_name, io_sheet_name):
        ws_sc_cat = self.wb[sc_cat_item_sheet_name]
        ws_io = self.wb[io_sheet_name]
        category_group = 'u_group_of_categories'
        top_category = 'u_top_category'
        code_col = 56
        field_name_col = 6
        value_col = 9
        sc_cat_io_row_number = 3
        sc_cat_io_row = ws_sc_cat[sc_cat_io_row_number]
        service_now_id =  sc_cat_io_row[code_col].value
        category_id = 00000000-0000-0000-0000-000000000000
        top_id = 00000000-0000-0000-0000-000000000000
        for i in range(3, ws_io.max_row + 1):
            row = ws_io[i]
            field = row[field_name_col].value
            if field == category_group:
                category_id = row[value_col].value
            if field == top_category:
                top_id = row[value_col].value
        with open('resources/source.txt', 'w', encoding='utf-8') as f:
            f.write(f'IteTile; {service_now_id}\nIteGroupCategory; {category_id}\nIteTopCategories; {top_id}')


    def fill_scripts_map (self, script_sheet_name):
        ws = self.wb[script_sheet_name]
        variable_name_col = 8
        script_name_col = 15
        script_action_col = 17 
        for i in range(3, ws.max_row + 1):
            row = ws[i]
            variable_name = row[variable_name_col].value
            if len(variable_name) > 1:
                variable_name = variable_name[3:]
            script_name = row[script_name_col].value
            script_action = row[script_action_col].value
            self.add_in_script_map(variable_name, script_name, script_action)


    def add_in_script_map(self, variable_name, script_name, script_action): 
        script_info = {
                'name': script_name,
                'action': script_action
            }
        if not self.scripts_map.keys().__contains__(variable_name):
            self.scripts_map[variable_name] = {
                'name': '',
                'script_list': []
            }
        self.scripts_map[variable_name]['script_list'].append(script_info)


    def try_to_fill_field_names(self, io_sheet_name):
        io_sheet = self.wb[io_sheet_name]
        field_id_col = 5
        field_name_col = 6
        for row_number in range(4, io_sheet.max_row + 1):
            row = io_sheet[row_number]
            var_id = row[field_id_col].value
            if var_id in self.scripts_map:
                field_name = row[field_name_col].value
                self.scripts_map[var_id]['name'] = field_name


    def write_in_actions_file(self):
        with open('out/scripts.txt', 'w', encoding='utf-8') as f:
            for variable in self.scripts_map.keys():
                var_name = self.scripts_map[variable]['name']
                scripts_of_variable = self.scripts_map[variable]['script_list']
                f.write(f'{var_name}\n{variable}\n')
                for script_info in scripts_of_variable:
                    name = script_info['name']
                    action = script_info['action']
                    f.write(f'{name}\n\n{action}\n\n-----------------------------------------------------\n')
                f.write(f'=====================================================\n')                    


    def print_rules_actions(self):
        for field in self.rules_action:
            print(f'{field}:')
            for rule in self.rules_action[field]['rules']:
                print(f'\t{rule}:')
                print(f"\t\t{self.rules_action[field]['rules'][rule]['rule']}")
                print(f"\t\t{self.rules_action[field]['rules'][rule]['visible']}")
                print(f"\t\t{self.rules_action[field]['rules'][rule]['required']}")
                print(f"\t\t{self.rules_action[field]['rules'][rule]['enabled']}")

    def execute(self, rule_action_sheet_name, rule_sheet_name, io_sheet_name, sc_cat_item_sheet_name, script_sheet_name):
        self.fill_id_to_names_list(io_sheet_name)
        self.fill_source_file(sc_cat_item_sheet_name, io_sheet_name)
        self.read_rules_actions(rule_action_sheet_name)
        self.read_rules(rule_sheet_name)
        self.try_fill_full_field_name(io_sheet_name)
        self.write_in_file()
        self.fill_scripts_map(script_sheet_name)
        self.try_to_fill_field_names(io_sheet_name)
        self.write_in_actions_file()

# sheet.max_row
# sheet.max_column
def main():
    excelparser = ExcelTileParser('./resources/выборка_Запрос на подтверждение или изменение информации об ИТ-активе (1).xlsx')
    # Для старых выгрузок
    # excelparser.execute(
    #     rule_action_sheet_name='catalog_ui_policy_action_дей.по', 
    #     rule_sheet_name='catalog_ui_policy_политики', 
    #     io_sheet_name='item_option_new', 
    #     sc_cat_item_sheet_name='sc_cat_item_плитка'
    #     )
    # Для новых выгрузок
    excelparser.execute(
        rule_action_sheet_name='действия политик', 
        rule_sheet_name='политики', 
        io_sheet_name='переменные', 
        sc_cat_item_sheet_name='sc_cat_item',
        script_sheet_name='скрипты'
        )

if __name__ == '__main__':
    main()